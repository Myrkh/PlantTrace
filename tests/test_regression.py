from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from planttrace.batch import load_reference_file, parse_references, run_batch_search

from planttrace.cli import main
from planttrace.export import export_batch, export_results
from planttrace.extractor import extract_references
from planttrace.models import ExtractionRule
from planttrace.normalization import compact_identifier, compact_ocr_identifier, exact_variants
from planttrace.rule_packs import export_rule_pack, import_rule_pack, load_preset_pack, project_rule_pack
from planttrace.rules import load_rules, load_stoplist, save_rules, save_stoplist
from planttrace.pdf_engine import render_page_preview
from planttrace.search import search
from planttrace.updates import is_newer
from tests.support import make_pdf, run_root


def test_normalization_covers_industrial_tag_variants() -> None:
    assert compact_identifier("FV-1100") == "FV1100"
    assert compact_ocr_identifier("FO-IIOO") == "F01100"
    assert {"FV-1100", "FV 1100", "FV_1100"}.issubset(exact_variants("FV1100"))


def test_rules_roundtrip_to_project_file() -> None:
    root = run_root()
    rules = [ExtractionRule(name="Custom tags", kind="TAG", pattern=r"\b[A-Z]{2}\d{4}\b")]

    save_rules(root, rules)
    loaded = load_rules(root)

    assert loaded == rules


def test_rule_pack_roundtrip_and_preset_loading() -> None:
    root = run_root()
    preset = load_preset_pack("Instrumentation")

    assert preset.rules
    assert any(rule.kind == "TAG" for rule in preset.rules)

    custom = project_rule_pack("Client A", [ExtractionRule(name="Client tags", kind="TAG", pattern=r"\bAA\d{4}\b")], {"RJ45"})
    output = root / "client-a-rule-pack.json"
    export_rule_pack(custom, output)
    imported = import_rule_pack(output)

    assert imported.name == "Client A"
    assert imported.rules == custom.rules
    assert imported.stoplist == {"RJ45"}


def test_stoplist_roundtrip_and_extraction_filtering() -> None:
    root = run_root()
    pdf_root = root / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "refs.pdf", ["RJ45 and FV-1100 are both visible."])
    main(["index", "--project", str(root), "--pdf-root", str(pdf_root)])

    save_stoplist(root, {"RJ45"})
    values = {hit.value for hit in extract_references(root)}

    assert load_stoplist(root) == {"RJ45"}
    assert "RJ45" not in values
    assert "FV-1100" in values


def test_batch_search_reads_reference_list_and_exports_matrix() -> None:
    root = run_root()
    pdf_root = root / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "refs.pdf", ["FV-1100 appears here. JDY checked the package."])
    main(["index", "--project", str(root), "--pdf-root", str(pdf_root)])

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["TAG"])
    sheet.append(["FV1100"])
    sheet.append(["JDY"])
    sheet.append(["ZZ9999"])
    reference_file = root / "references.xlsx"
    workbook.save(reference_file)

    references = load_reference_file(reference_file)
    results = run_batch_search(root, references, "hybrid")
    output = root / "batch.xlsx"
    export_batch(results, output)

    assert references == ["FV1100", "JDY", "ZZ9999"]
    assert parse_references("FV1100; JDY\nFV1100") == ["FV1100", "JDY"]
    assert [result.status for result in results] == ["found", "found", "absent_indexed_text"]
    assert output.exists()
    assert load_workbook(output).active["A1"].value == "reference"


def test_cli_index_search_coverage_and_extract(capsys: object) -> None:
    root = run_root()
    pdf_root = root / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "loop.pdf", ["FV-1100 and line 10-P-12345 are visible."])

    assert main(["index", "--project", str(root), "--pdf-root", str(pdf_root)]) == 0
    assert main(["search", "--project", str(root), "--query", "FV1100", "--mode", "exact"]) == 0
    assert main(["coverage", "--project", str(root)]) == 0
    assert main(["extract", "--project", str(root), "--limit", "10"]) == 0

    output = capsys.readouterr().out
    assert "exact_normalized" in output
    assert "FV-1100" in output
    assert "10-P-12345" in output


def test_export_results_xlsx_has_expected_headers() -> None:
    root = run_root()
    pdf_root = root / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "loop.pdf", ["FV-1100 appears here."])
    main(["index", "--project", str(root), "--pdf-root", str(pdf_root)])
    results = search(root, "FV1100", mode="exact")

    output = root / "results.xlsx"
    export_results(results, output)

    workbook = load_workbook(output)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]][:4] == ["query", "match_type", "document_path", "page"]


def test_render_page_preview_returns_png_and_highlights() -> None:
    root = run_root()
    pdf = root / "loop.pdf"
    make_pdf(pdf, ["FV-1100 appears here.", "nothing relevant on this page"])

    png, rects = render_page_preview(pdf, 1, ["FV-1100"])
    assert png[:8] == b"\x89PNG\r\n\x1a\n"
    assert len(rects) >= 1

    empty_png, empty_rects = render_page_preview(pdf, 2, ["FV-1100"])
    assert empty_png[:8] == b"\x89PNG\r\n\x1a\n"
    assert empty_rects == []


def test_reference_completeness_flags_present_and_missing() -> None:
    from planttrace.reference_profile import ReferenceOccurrence, compute_completeness

    occurrences = [
        ReferenceOccurrence(
            family="Loop diagram", document_path="a.pdf", filename="a.pdf", page=1,
            match_type="exact_normalized", found_as="FV-1100", excerpt="", page_status="ok", document_status="ok",
        )
    ]
    result = dict(compute_completeness(occurrences))
    assert result["Loop diagram"] is True
    assert result["Datasheet"] is False
    assert result["PID / PFD"] is False


def test_bug_report_builds_diagnostic_and_subject() -> None:
    from planttrace import __version__
    from planttrace.bug_report import diagnostic_text, report_body, subject

    assert __version__ in subject()
    diagnostic = diagnostic_text("C:/Projet", ocr_detected=False, semantic="offline")
    assert "Version : " in diagnostic
    assert "OCR Tesseract : non detecte" in diagnostic
    assert "C:/Projet" in diagnostic
    body = report_body(diagnostic)
    assert body.endswith(diagnostic)
    assert "Decrivez le probleme" in body


def test_self_update_helper_and_locate() -> None:
    from planttrace import self_update

    staging = run_root()
    app_dir = staging / "PlantTrace"
    app_dir.mkdir(parents=True)
    (app_dir / "PlantTrace.exe").write_bytes(b"x")

    assert self_update._locate_app_dir(staging) == app_dir
    assert self_update.is_supported() is False

    helper = self_update._write_helper(staging, 4242, app_dir, Path("C:/Apps/PlantTrace"), Path("C:/Apps/PlantTrace/PlantTrace.exe"))
    text = helper.read_text(encoding="utf-8")
    assert helper.name == "apply_update.bat"
    assert "4242" in text
    assert "robocopy" in text
    assert "PlantTrace.exe" in text


def test_changelog_is_well_formed() -> None:
    from planttrace.changelog import RELEASES

    assert RELEASES
    for release in RELEASES:
        assert release.version and release.date and release.tagline
        assert release.sections
        for section in release.sections:
            assert section.title and section.items
            assert all(isinstance(item, str) and item for item in section.items)


def test_is_newer_compares_versions_numerically() -> None:
    assert is_newer("0.2.0", "0.1.0")
    assert is_newer("0.10.0", "0.9.0")
    assert is_newer("v0.1.1", "0.1.0")
    assert not is_newer("0.1.0", "0.1.0")
    assert not is_newer("0.1.0", "0.2.0")


def test_runtime_has_no_network_imports() -> None:
    source_root = Path("src") / "planttrace"
    forbidden = ("requests", "httpx", "urllib", "socket")
    # updates.py is the single sanctioned network module: an explicit, user-initiated
    # check of the latest GitHub Release. The evidence pipeline (index/search/extract)
    # must stay offline, so every other module is still scanned.
    allowed = {"updates.py"}
    offenders: list[str] = []
    for path in source_root.rglob("*.py"):
        if path.name in allowed:
            continue
        text = path.read_text(encoding="utf-8")
        for name in forbidden:
            if f"import {name}" in text or f"from {name}" in text:
                offenders.append(f"{path}:{name}")

    assert offenders == []
