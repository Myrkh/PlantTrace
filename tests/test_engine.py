from __future__ import annotations

import json
from zipfile import ZipFile

import fitz

from planttrace.batch import run_batch_search
from planttrace.conflicts import detect_conflicts
from planttrace.deliverable import build_deliverable_pack
from planttrace.doc_families import DocumentFamily, classify_documents
from planttrace.indexer import index_folder
from planttrace.master_register import MasterRegisterConfig, build_master_register
from planttrace.coverage import coverage_documents, coverage_tasks, coverage_triage, coverage_verdict
from planttrace.extractor import extract_references
from planttrace.export import export_conflicts, export_coverage, export_coverage_tasks, export_doc_families, export_extraction, export_revisions
from planttrace.matrix_exports import export_project_matrix
from planttrace.project_matrix import ProjectMatrixRow, build_project_matrix
from planttrace.reference_exports import export_reference_profile
from planttrace.reference_profile import build_reference_profile
from planttrace.revisions import RevisionChange, compare_revision_folders
from planttrace.search import search
from planttrace.store import PlantTraceStore
from planttrace.template_exports import export_template_run
from planttrace.templates import TAG_REGISTER_TEMPLATE, TemplateRun, run_template
from planttrace.models import ProjectPaths
from tests.support import make_pdf, run_root


def make_template(path, sheet_name: str, headers: list[str]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    workbook.save(path)


def make_source_list(path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LISTE"
    sheet.append(["TAG", "SERVICE", "FONCTION", "PID", "DATA SHEET", "SYSTEME_ACT", "SYSTEME_FUT", "JB_TAG"])
    sheet.append(["520FT1101", "CHARGE PTT UCO", "DELTA P", "HTI199-020-PRO-3304-PID", "HTI199-020-INS-2235-SPE", "", "UCN16 HPM11/12", "520BJM1"])
    sheet.append(["340G1012_E", "340G1012 ENCLENCHEMENT", "CDE TOR", "HTI199-020-PRO-3308-PID", "", "410AUT30", "410AUT38", "P2-ST10-100"])
    workbook.save(path)


def test_exact_normalized_search_finds_tag_variants() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "loop.pdf", ["Loop drawing references FV-1100 and line 10-P-12345."])

    report = index_folder(tmp_path, pdf_root)

    assert report.indexed == 1
    results = search(tmp_path, "FV1100", mode="hybrid")
    assert results[0].match_type == "exact_normalized"
    assert results[0].page == 1
    assert "FV-1100" in results[0].excerpt


def test_index_reports_progress_per_file() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    for name in ("a.pdf", "b.pdf", "c.pdf"):
        make_pdf(pdf_root / name, [f"Document {name} mentions FV-1100."])

    events: list[tuple[int, int, str]] = []
    index_folder(tmp_path, pdf_root, progress_callback=lambda done, total, name: events.append((done, total, name)))

    assert all(total == 3 for _done, total, _name in events)
    assert [done for done, _total, _name in events] == [0, 1, 2, 3]
    assert events[-1] == (3, 3, "")


def test_text_search_and_not_found_are_evidence_shaped() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "vendor.pdf", ["Le fournisseur de cafe est mentionne dans cette note."])
    index_folder(tmp_path, pdf_root)

    text_results = search(tmp_path, "fournisseur cafe", mode="text")
    assert text_results
    assert text_results[0].document_path.endswith("vendor.pdf")
    assert text_results[0].page == 1

    missing = search(tmp_path, "ZZ9999", mode="exact")
    assert missing[0].match_type == "not_found_in_indexed_text"


def test_fuzzy_search_handles_imprecise_words() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "note.pdf", ["Le fournisseur de cafe industriel est ACME Process."])
    index_folder(tmp_path, pdf_root)

    results = search(tmp_path, "fourniseur cafee", mode="fuzzy")

    assert results
    assert results[0].match_type == "fuzzy_text"
    assert results[0].document_path.endswith("note.pdf")


def test_coverage_tracks_ocr_required_pages() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    document = fitz.open()
    document.new_page()
    page = document.new_page()
    page.insert_text((72, 72), "HS 0900 visible")
    document.save(pdf_root / "mixed.pdf")
    document.close()

    index_folder(tmp_path, pdf_root)
    coverage = PlantTraceStore(ProjectPaths(tmp_path)).coverage()

    assert coverage["documents"] == 1
    assert coverage["pages"] == 2
    assert coverage["text_pages"] == 1
    assert coverage["ocr_required_pages"] == 1
    assert "incomplete" in coverage_verdict(coverage)
    documents = coverage_documents(tmp_path)
    assert documents[0].coverage_percent == 50.0
    assert documents[0].triage == "ocr_required"
    assert coverage_triage(2, 2, 0, 0) == "ok"
    tasks = coverage_tasks(tmp_path)
    assert tasks[0].filename == "mixed.pdf"
    assert tasks[0].page == 1
    assert tasks[0].action == "Indexer avec OCR"

    output = tmp_path / "coverage.xlsx"
    export_coverage(documents, output)
    assert output.exists()
    triage_output = tmp_path / "ocr-triage.xlsx"
    export_coverage_tasks(tasks, triage_output)
    assert triage_output.exists()


def test_extraction_finds_project_references_and_exports() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "refs.pdf", ["FV-1100 is connected to line 10-P-12345 in document HTI199-VEN-ELE-3799-300139-01."])
    index_folder(tmp_path, pdf_root)

    hits = extract_references(tmp_path)
    values = {hit.value for hit in hits}

    assert "FV-1100" in values
    assert "10-P-12345" in values
    assert any(hit.kind == "DOC" for hit in hits)

    output = tmp_path / "extraction.xlsx"
    export_extraction(hits, output)

    assert output.exists()


def test_conflict_detector_finds_conflicting_lines_for_same_tag() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "pid.pdf", ["FV-1100 is connected to line 10-P-12345."])
    make_pdf(pdf_root / "loop.pdf", ["FV-1100 is connected to line 10-P-99999."])
    index_folder(tmp_path, pdf_root)

    findings = detect_conflicts(tmp_path)
    output = tmp_path / "conflicts.xlsx"
    export_conflicts(findings, output)

    assert output.exists()
    assert findings
    assert findings[0].reference == "FV-1100"
    assert findings[0].field == "LINE"
    assert findings[0].severity == "high"
    assert "10-P-12345" in findings[0].values
    assert "10-P-99999" in findings[0].values


def test_document_family_classifier_labels_indexed_pdf_types() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "PID-001.pdf", ["Piping and instrumentation diagram for FV-1100."])
    make_pdf(pdf_root / "loop-001.pdf", ["Loop diagram for FV-1100."])
    make_pdf(pdf_root / "vendor.pdf", ["Installation manual from supplier ACME."])
    index_folder(tmp_path, pdf_root)

    families = classify_documents(tmp_path)
    output = tmp_path / "families.xlsx"
    export_doc_families(families, output)

    by_file = {family.filename: family.family for family in families}
    assert output.exists()
    assert by_file["PID-001.pdf"] == "PID_PFD"
    assert by_file["loop-001.pdf"] == "LOOP"
    assert by_file["vendor.pdf"] == "VENDOR"


def test_reference_profile_consolidates_occurrences_associations_and_alerts() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "pid.pdf", ["Piping and instrumentation diagram. FV-1100 is connected to line 10-P-12345."])
    make_pdf(pdf_root / "loop.pdf", ["Loop diagram. FV-1100 is connected to line 10-P-99999."])
    index_folder(tmp_path, pdf_root)

    conflicts = detect_conflicts(tmp_path)
    profile = build_reference_profile(tmp_path, "FV1100", conflicts, [])
    output = tmp_path / "reference.xlsx"
    export_reference_profile(profile, output)

    assert output.exists()
    assert profile.occurrence_count >= 2
    assert profile.document_count == 2
    assert any(association.kind == "LINE" and "10-P-12345" in association.value for association in profile.associations)
    assert any(alert.source == "Conflits" and alert.field == "LINE" for alert in profile.alerts)


def test_project_matrix_crosses_references_with_families_and_alerts() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "pid.pdf", ["Piping and instrumentation diagram. FV-1100 is connected to line 10-P-12345."])
    make_pdf(pdf_root / "loop.pdf", ["Loop diagram. FV-1100 is connected to line 10-P-99999."])
    index_folder(tmp_path, pdf_root)

    conflicts = detect_conflicts(tmp_path)
    rows = build_project_matrix(tmp_path, conflicts, [])
    output = tmp_path / "matrix.xlsx"
    export_project_matrix(rows, output)

    tag_row = next(row for row in rows if row.reference == "FV-1100")
    assert output.exists()
    assert tag_row.pid_pfd > 0
    assert tag_row.loop > 0
    assert tag_row.conflicts == "high:LINE"


def test_tag_register_template_fills_engineering_columns_with_evidence() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "pid.pdf", ["Piping and instrumentation diagram. FV-1100 description: feed valve connected to line 10-P-12345."])
    make_pdf(pdf_root / "loop.pdf", ["Loop diagram. FV-1100 is connected to line 10-P-99999 in document HTI199-VEN-ELE-3799-300139-01."])
    index_folder(tmp_path, pdf_root)

    conflicts = detect_conflicts(tmp_path)
    run = run_template(tmp_path, TAG_REGISTER_TEMPLATE, conflicts, [])
    output = tmp_path / "tag-register.xlsx"
    export_template_run(run, output)

    row = next(item for item in run.rows if item.tag == "FV-1100")
    assert output.exists()
    assert "10-P-12345" in row.lines
    assert "10-P-99999" in row.lines
    assert "HTI199" in row.documents
    assert "high:LINE" in row.conflicts


def test_master_register_imports_client_templates_and_exports_mapped_registers() -> None:
    tmp_path = run_root()
    source_root = tmp_path / "source"
    source_root.mkdir()
    pdf_root = source_root / "pdfs"
    pdf_root.mkdir()
    make_source_list(source_root / "HTI199-020-INS-0002-LST_04.xlsx")
    make_pdf(pdf_root / "HTI199-020-INS-2810-WIR_01.pdf", ["Wiring diagram references 520FT1101 and terminal 520BJM1."])
    index_folder(tmp_path, pdf_root)

    tags_template = tmp_path / "Tags Template.xlsx"
    links_template = tmp_path / "Tags Documents Links Template.xlsx"
    make_template(tags_template, "Tags List", ["PlantNo", "Site", "TagNo", "Description", "TagType", "Sector", "System", "SubSystem", "Class", "CommunicationTag", "Deleted"])
    make_template(links_template, "Tags Documents Links List", ["PlantNo", "TagNo", "Site", "DocumentID", "Deleted"])

    result = build_master_register(
        MasterRegisterConfig(
            source_root=source_root,
            tags_template=tags_template,
            links_template=links_template,
            output_dir=tmp_path / "out",
            plant_no="HTI199",
            site="GPS",
        ),
        project_root=tmp_path,
    )

    from openpyxl import load_workbook

    tags_wb = load_workbook(result.tags_output, data_only=True)
    links_wb = load_workbook(result.links_output, data_only=True)
    tags_sheet = tags_wb["Tags List"]
    links_sheet = links_wb["Tags Documents Links List"]

    assert result.tags_output.exists()
    assert result.links_output.exists()
    assert result.evidence_output.exists()
    assert result.link_count >= 5
    tag_rows = {row[2]: row for row in tags_sheet.iter_rows(min_row=2, values_only=True) if row[2]}
    assert tag_rows["520FT1101"][0] == "HTI199"
    assert tag_rows["520FT1101"][1] == "GPS"
    assert tag_rows["520FT1101"][3] == "CHARGE PTT UCO"
    assert tag_rows["520FT1101"][4] == "DELTA P"
    assert tag_rows["520FT1101"][6] == "UCN16 HPM11/12"
    links = {(row[1], row[3]) for row in links_sheet.iter_rows(min_row=2, values_only=True) if row[1]}
    assert ("520FT1101", "HTI199-020-PRO-3304-PID") in links
    assert ("520FT1101", "HTI199-020-INS-2235-SPE") in links
    assert ("520FT1101", "HTI199-020-INS-2810-WIR") in links

    pack_output = tmp_path / "pack.zip"
    pack = build_deliverable_pack(tmp_path, pack_output, [], [], [], master_register=result)
    with ZipFile(pack.output) as archive:
        names = set(archive.namelist())
        manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
    assert "exports/master-register/Tags Template - PlantTrace.xlsx" in names
    assert "exports/master-register/Tags Documents Links Template - PlantTrace.xlsx" in names
    assert "exports/master-register/PlantTrace Master Register Evidence.xlsx" in names
    assert manifest["counts"]["master_register_tags"] == len(result.tags)
    assert manifest["counts"]["master_register_links"] == result.link_count


def test_revision_compare_finds_added_removed_and_modified_references() -> None:
    tmp_path = run_root()
    old_root = tmp_path / "old"
    new_root = tmp_path / "new"
    old_root.mkdir()
    new_root.mkdir()
    make_pdf(old_root / "pid.pdf", ["FV-1100 is connected to line 10-P-12345. PT-2000 remains old only."])
    make_pdf(new_root / "pid.pdf", ["FV-1100 is connected to line 10-P-99999. XV-3000 is new only."])

    comparison = compare_revision_folders(tmp_path, old_root, new_root)
    output = tmp_path / "revisions.xlsx"
    export_revisions(comparison.changes, output)

    changes = {(change.status, change.kind, change.reference) for change in comparison.changes}
    assert output.exists()
    assert ("modified", "TAG", "FV-1100") in changes
    assert ("removed", "TAG", "PT-2000") in changes
    assert ("added", "TAG", "XV-3000") in changes


def test_deliverable_pack_contains_audit_manifest_and_available_exports() -> None:
    tmp_path = run_root()
    pdf_root = tmp_path / "pdfs"
    pdf_root.mkdir()
    make_pdf(pdf_root / "refs.pdf", ["FV-1100 is connected to line 10-P-12345."])
    index_folder(tmp_path, pdf_root)

    search_results = search(tmp_path, "FV1100", mode="hybrid")
    batch_results = run_batch_search(tmp_path, ["FV1100", "ZZ9999"])
    extraction_hits = extract_references(tmp_path)
    revision_changes = [
        RevisionChange("added", "TAG", "XV-3000", 0, 1, "", "pid.pdf p1", "", "XV-3000", "Reference presente en nouvelle revision uniquement.")
    ]
    doc_families = [
        DocumentFamily("LOOP", "Loop diagram", "high", 7, "loop.pdf", "loop.pdf", 1, "texte: LOOP DIAGRAM", "ok")
    ]
    reference_profile = build_reference_profile(tmp_path, "FV1100", [], [])
    project_matrix = [
        ProjectMatrixRow("TAG", "FV-1100", 1, 1, "Loop", 0, 1, 0, 0, 0, 0, 0, 0, 0, "loop.pdf", "loop.pdf p1", "", "", "FV-1100")
    ]
    template_run = TemplateRun(
        TAG_REGISTER_TEMPLATE,
        run_template(tmp_path, TAG_REGISTER_TEMPLATE, [], []).rows,
    )
    output = tmp_path / "planttrace-livrable.zip"

    pack = build_deliverable_pack(
        tmp_path,
        output,
        search_results,
        batch_results,
        extraction_hits,
        revision_changes=revision_changes,
        doc_families=doc_families,
        reference_profile=reference_profile,
        project_matrix=project_matrix,
        template_run=template_run,
    )

    assert pack.output == output
    assert output.exists()
    with ZipFile(output) as archive:
        names = set(archive.namelist())
        assert "manifest.json" in names
        assert "README.txt" in names
        assert "exports/search-results.xlsx" in names
        assert "exports/batch-matrix.xlsx" in names
        assert "exports/extraction-inventory.xlsx" in names
        assert "exports/reference-profile.xlsx" in names
        assert "exports/project-matrix.xlsx" in names
        assert "exports/tag-register-template.xlsx" in names
        assert "exports/document-families.xlsx" in names
        assert "exports/revision-compare.xlsx" in names
        assert "exports/coverage-report.xlsx" in names
        assert "audit/absences-qualifiees.csv" in names
        manifest = json.loads(archive.read("manifest.json").decode("utf-8"))

    assert manifest["counts"]["batch_rows"] == 2
    assert manifest["counts"]["reference_profile"] == 1
    assert manifest["counts"]["project_matrix_rows"] == 1
    assert manifest["counts"]["template_rows"] == len(template_run.rows)
    assert manifest["counts"]["doc_families"] == 1
    assert manifest["counts"]["revision_changes"] == 1
    assert manifest["counts"]["coverage_documents"] == 1
    assert "Absence" not in manifest["coverage_verdict"]
