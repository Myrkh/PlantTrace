from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import ProjectPaths
from .normalization import compact_identifier
from .store import PlantTraceStore


TAG_LIST_SHEET = "Tags List"
TAG_LINKS_SHEET = "Tags Documents Links List"

TAG_HEADERS = [
    "PlantNo",
    "Site",
    "TagNo",
    "Description",
    "TagType",
    "Sector",
    "System",
    "SubSystem",
    "Class",
    "CommunicationTag",
    "Deleted",
]

LINK_HEADERS = ["PlantNo", "TagNo", "Site", "DocumentID", "Deleted"]


@dataclass(frozen=True)
class MasterRegisterConfig:
    source_root: Path
    tags_template: Path
    links_template: Path
    output_dir: Path
    plant_no: str
    site: str


@dataclass
class TagEvidence:
    source: str
    location: str
    field: str
    value: str


@dataclass
class MasterTag:
    tag_no: str
    plant_no: str
    site: str
    description: str = ""
    tag_type: str = ""
    sector: str = ""
    system: str = ""
    sub_system: str = ""
    tag_class: str = ""
    communication_tag: str = ""
    deleted: str = ""
    documents: set[str] = field(default_factory=set)
    evidence: list[TagEvidence] = field(default_factory=list)


@dataclass(frozen=True)
class MasterRegisterResult:
    tags: list[MasterTag]
    link_count: int
    tags_output: Path
    links_output: Path
    evidence_output: Path


def build_master_register(config: MasterRegisterConfig, project_root: Path | None = None) -> MasterRegisterResult:
    tags = collect_source_tags(config)
    if project_root is not None:
        add_indexed_pdf_links(project_root, tags)
    sorted_tags = sorted(tags.values(), key=lambda item: item.tag_no)
    config.output_dir.mkdir(parents=True, exist_ok=True)

    tags_output = config.output_dir / "Tags Template - PlantTrace.xlsx"
    links_output = config.output_dir / "Tags Documents Links Template - PlantTrace.xlsx"
    evidence_output = config.output_dir / "PlantTrace Master Register Evidence.xlsx"
    export_tags_template(config.tags_template, tags_output, sorted_tags)
    link_count = export_links_template(config.links_template, links_output, sorted_tags)
    export_evidence(evidence_output, sorted_tags)

    return MasterRegisterResult(sorted_tags, link_count, tags_output, links_output, evidence_output)


def collect_source_tags(config: MasterRegisterConfig) -> dict[str, MasterTag]:
    tags: dict[str, MasterTag] = {}
    for workbook_path in source_workbooks(config.source_root, {config.tags_template, config.links_template}):
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            header_row, headers = detect_header(sheet)
            if not headers or "TAG" not in headers.values():
                continue
            for excel_row in sheet.iter_rows(min_row=header_row + 1, max_col=max(headers) + 1, values_only=True):
                values = row_values(headers, excel_row)
                tag_no = clean_cell(values.get("TAG", ""))
                if not is_probable_tag(tag_no):
                    continue
                tag = tags.setdefault(tag_no, MasterTag(tag_no=tag_no, plant_no=config.plant_no, site=config.site))
                merge_tag(tag, values, workbook_path, sheet.title, header_row)
    return tags


def source_workbooks(source_root: Path, excluded: set[Path]) -> Iterable[Path]:
    excluded_resolved = {path.resolve() for path in excluded if path.exists()}
    for path in sorted(Path(source_root).rglob("*.xlsx")):
        if path.resolve() in excluded_resolved:
            continue
        if path.name.startswith("~$"):
            continue
        yield path


def detect_header(sheet: Worksheet, max_scan_rows: int = 30) -> tuple[int, dict[int, str]]:
    best_row = 0
    best_headers: dict[int, str] = {}
    best_score = 0
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        headers = {index: canonical_header(value) for index, value in enumerate(row) if canonical_header(value)}
        score = sum(1 for value in headers.values() if value in {"TAG", "DESCRIPTION", "TAGTYPE", "SYSTEM", "SUBSYSTEM", "PID", "DATASHEET", "HOOKUP", "DOCUMENTID"})
        if "TAG" in headers.values():
            score += 5
        if score > best_score:
            best_row = row_index
            best_headers = headers
            best_score = score
    if "TAG" not in best_headers.values():
        return 0, {}
    return best_row, best_headers


def canonical_header(value: object) -> str:
    text = clean_cell(value)
    key = re.sub(r"[^A-Z0-9]+", "", text.upper())
    aliases = {
        "TAG": "TAG",
        "TAGNO": "TAG",
        "TAGNUMBER": "TAG",
        "TAGDESCRIPTION": "DESCRIPTION",
        "DESCRIPTION": "DESCRIPTION",
        "SERVICE": "DESCRIPTION",
        "TAGTYPE": "TAGTYPE",
        "FUNCTION": "TAGTYPE",
        "FONCTION": "TAGTYPE",
        "SYSTEM": "SYSTEM",
        "SYSTEMEACT": "SYSTEM",
        "SYSTEMEFUT": "SYSTEM",
        "TAGSYSTEM": "SYSTEM",
        "SUBSYSTEM": "SUBSYSTEM",
        "SUBFUNCTION": "SUBSYSTEM",
        "SUBFUNCTION": "SUBSYSTEM",
        "TAGSUBSYSTEM": "SUBSYSTEM",
        "SECTOR": "SECTOR",
        "CLASS": "CLASS",
        "PLANTNUMBER": "SECTOR",
        "PID": "PID",
        "PANDID": "PID",
        "DATASHEET": "DATASHEET",
        "HOOKUP": "HOOKUP",
        "DOCUMENTID": "DOCUMENTID",
        "DOCUMENT": "DOCUMENTID",
        "COMMUNICATIONTAG": "COMMUNICATIONTAG",
        "DELETED": "DELETED",
        "ACTION": "ACTION",
    }
    return aliases.get(key, "")


def row_values(headers: dict[int, str], row: tuple[object, ...]) -> dict[str, str]:
    values: dict[str, str] = {}
    for index, header in headers.items():
        if not header:
            continue
        value = clean_cell(row[index] if index < len(row) else "")
        if value and not values.get(header):
            values[header] = value
        else:
            values.setdefault(header, "")
    return values


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return " ".join(text.split())


def is_probable_tag(value: str) -> bool:
    if not value or len(value) < 3:
        return False
    if value.upper() in {"TAG", "N/A", "NA", "NONE"}:
        return False
    return bool(re.search(r"\d", value)) and bool(re.search(r"[A-Za-z]", value))


def merge_tag(tag: MasterTag, values: dict[str, str], workbook_path: Path, sheet_name: str, header_row: int) -> None:
    apply_first(tag, "description", values.get("DESCRIPTION", ""))
    apply_first(tag, "tag_type", values.get("TAGTYPE", ""))
    apply_first(tag, "sector", values.get("SECTOR", ""))
    apply_first(tag, "system", values.get("SYSTEM", ""))
    apply_first(tag, "sub_system", values.get("SUBSYSTEM", ""))
    apply_first(tag, "tag_class", values.get("CLASS", ""))
    apply_first(tag, "communication_tag", values.get("COMMUNICATIONTAG", ""))
    if values.get("DELETED"):
        tag.deleted = values["DELETED"]
    elif values.get("ACTION", "").upper() in {"SUP", "SUPPRIME", "DELETED", "DESAFFECTE"}:
        tag.deleted = "oui"

    source_doc = document_id_from_filename(workbook_path.name)
    add_document(tag, source_doc, workbook_path.name, sheet_name, "source_workbook")
    for field in ["PID", "DATASHEET", "HOOKUP", "DOCUMENTID"]:
        for document_id in split_document_ids(values.get(field, "")):
            add_document(tag, document_id, workbook_path.name, sheet_name, field)
    tag.evidence.append(TagEvidence(workbook_path.name, f"{sheet_name} header row {header_row}", "TAG", tag.tag_no))


def apply_first(tag: MasterTag, field_name: str, value: str) -> None:
    if value and not getattr(tag, field_name):
        setattr(tag, field_name, value)


def split_document_ids(value: str) -> list[str]:
    if not value:
        return []
    parts = re.split(r"\s*(?:\||;|,|\n)\s*", value)
    return [clean_document_id(part) for part in parts if clean_document_id(part)]


def clean_document_id(value: str) -> str:
    text = clean_cell(value)
    if not text or text.upper() in {"N/A", "NA", "PACKAGE"}:
        return ""
    if re.match(r"^[A-Z]{2,}\d{3}[-_]", text, flags=re.IGNORECASE):
        text = document_id_from_filename(text)
    return text


def document_id_from_filename(filename: str) -> str:
    stem = Path(filename).stem.strip(" _-")
    stem = re.split(r"\s+-\s+", stem, maxsplit=1)[0].strip(" _-")
    stem = re.sub(r"\s*\([^)]*\)$", "", stem).strip(" _-")
    stem = re.sub(r"[_ -]+R?\d{1,2}(?:\s+[A-Z].*)?$", "", stem, flags=re.IGNORECASE).strip(" _-")
    return stem


def add_document(tag: MasterTag, document_id: str, source: str, location: str, field_name: str) -> None:
    clean = clean_document_id(document_id)
    if not clean:
        return
    tag.documents.add(clean)
    tag.evidence.append(TagEvidence(source, location, field_name, clean))


def add_indexed_pdf_links(project_root: Path, tags: dict[str, MasterTag]) -> None:
    if not ProjectPaths(project_root).db_path.exists():
        return
    pages = PlantTraceStore(ProjectPaths(project_root)).indexed_pages()
    compact_tags = {compact_identifier(tag_no): tag for tag_no, tag in tags.items()}
    compact_tags = {key: value for key, value in compact_tags.items() if key}
    if not compact_tags:
        return
    tag_pattern = re.compile("|".join(re.escape(key) for key in sorted(compact_tags, key=len, reverse=True)))
    for page in pages:
        text = compact_identifier(page["text"])
        if not text:
            continue
        document_id = document_id_from_filename(Path(page["path"]).name)
        for compact_tag in {match.group(0) for match in tag_pattern.finditer(text)}:
            add_document(compact_tags[compact_tag], document_id, Path(page["path"]).name, f"PDF page {page['page_number']}", "PDF_OCCURRENCE")


def export_tags_template(template: Path, output: Path, tags: list[MasterTag]) -> None:
    workbook, sheet = workbook_sheet(template, TAG_LIST_SHEET)
    headers = sheet_headers(sheet, TAG_HEADERS)
    clear_data_rows(sheet, max(headers.values()))
    for row_index, tag in enumerate(tags, start=2):
        values = tag_row(tag)
        write_mapped_row(sheet, headers, row_index, values)
    workbook.save(output)


def export_links_template(template: Path, output: Path, tags: list[MasterTag]) -> int:
    workbook, sheet = workbook_sheet(template, TAG_LINKS_SHEET)
    headers = sheet_headers(sheet, LINK_HEADERS)
    clear_data_rows(sheet, max(headers.values()))
    row_index = 2
    for tag in tags:
        for document_id in sorted(tag.documents):
            write_mapped_row(
                sheet,
                headers,
                row_index,
                {
                    "PlantNo": tag.plant_no,
                    "TagNo": tag.tag_no,
                    "Site": tag.site,
                    "DocumentID": document_id,
                    "Deleted": tag.deleted,
                },
            )
            row_index += 1
    workbook.save(output)
    return row_index - 2


def workbook_sheet(template: Path, preferred_sheet: str) -> tuple[object, Worksheet]:
    workbook = load_workbook(template)
    if preferred_sheet in workbook.sheetnames:
        return workbook, workbook[preferred_sheet]
    return workbook, workbook.active


def clear_data_rows(sheet: Worksheet, max_column: int) -> None:
    for row in sheet.iter_rows(min_row=2, max_row=max(sheet.max_row, 2), max_col=max_column):
        for cell in row:
            cell.value = None


def sheet_headers(sheet: Worksheet, fallback: list[str]) -> dict[str, int]:
    headers = {clean_cell(cell.value): cell.column for cell in sheet[1] if clean_cell(cell.value)}
    if headers:
        return headers
    for index, header in enumerate(fallback, start=1):
        sheet.cell(row=1, column=index, value=header)
    return {header: index for index, header in enumerate(fallback, start=1)}


def tag_row(tag: MasterTag) -> dict[str, str]:
    return {
        "PlantNo": tag.plant_no,
        "Site": tag.site,
        "TagNo": tag.tag_no,
        "Description": tag.description,
        "TagType": tag.tag_type,
        "Sector": tag.sector,
        "System": tag.system,
        "SubSystem": tag.sub_system,
        "Class": tag.tag_class,
        "CommunicationTag": tag.communication_tag,
        "Deleted": tag.deleted,
    }


def write_mapped_row(sheet: Worksheet, headers: dict[str, int], row_index: int, values: dict[str, object]) -> None:
    for header, value in values.items():
        column = headers.get(header)
        if column:
            sheet.cell(row=row_index, column=column, value=value)


def export_evidence(output: Path, tags: list[MasterTag]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    tags_sheet = workbook.active
    tags_sheet.title = "summary"
    tags_sheet.append(["TagNo", "Description", "TagType", "Documents", "EvidenceCount"])
    for tag in tags:
        tags_sheet.append([tag.tag_no, tag.description, tag.tag_type, " | ".join(sorted(tag.documents)), len(tag.evidence)])
    evidence_sheet = workbook.create_sheet("evidence")
    evidence_sheet.append(["TagNo", "Source", "Location", "Field", "Value"])
    for tag in tags:
        for evidence in tag.evidence:
            evidence_sheet.append([tag.tag_no, evidence.source, evidence.location, evidence.field, evidence.value])
    workbook.save(output)
