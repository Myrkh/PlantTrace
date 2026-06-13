from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from .reference_profile import ReferenceProfile


def export_reference_profile(profile: ReferenceProfile, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = reference_profile_workbook(profile)
    workbook.save(output)
    workbook.close()


def reference_profile_xlsx_bytes(profile: ReferenceProfile) -> bytes:
    workbook = reference_profile_workbook(profile)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def reference_profile_workbook(profile: ReferenceProfile) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    summary.append(["field", "value"])
    summary.append(["query", profile.query])
    summary.append(["occurrences", profile.occurrence_count])
    summary.append(["documents", profile.document_count])
    summary.append(["families", profile.family_summary])
    summary.append(["associations", profile.association_count])
    summary.append(["alerts", profile.alert_count])

    occurrences = workbook.create_sheet("occurrences")
    occurrences.append(["family", "document_path", "filename", "page", "match_type", "found_as", "excerpt", "page_status", "document_status"])
    for item in profile.occurrences:
        occurrences.append([item.family, item.document_path, item.filename, item.page, item.match_type, item.found_as, item.excerpt, item.page_status, item.document_status])

    associations = workbook.create_sheet("associations")
    associations.append(["kind", "value", "evidence_count", "documents", "pages", "excerpt"])
    for item in profile.associations:
        associations.append([item.kind, item.value, item.evidence_count, item.documents, item.pages, item.excerpt])

    alerts = workbook.create_sheet("alerts")
    alerts.append(["source", "severity", "field", "values", "documents", "pages", "summary"])
    for item in profile.alerts:
        alerts.append([item.source, item.severity, item.field, item.values, item.documents, item.pages, item.summary])
    return workbook
