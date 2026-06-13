from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

from .models import SearchResult
from .search import search


@dataclass(frozen=True)
class BatchReferenceResult:
    reference: str
    status: str
    hit_count: int
    documents: str
    pages: str
    best_match_type: str
    best_excerpt: str


def parse_references(text: str) -> list[str]:
    values: list[str] = []
    for line in text.splitlines():
        for item in line.replace(";", ",").split(","):
            value = item.strip()
            if value:
                values.append(value)
    return dedupe(values)


def load_reference_file(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return load_xlsx_references(path)
    if suffix == ".csv":
        return load_csv_references(path)
    return parse_references(path.read_text(encoding="utf-8-sig"))


def load_csv_references(path: Path) -> list[str]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    return references_from_rows(rows)


def load_xlsx_references(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return references_from_rows(rows)


def references_from_rows(rows: list[list[object]]) -> list[str]:
    if not rows:
        return []
    column = preferred_column(rows[0])
    values = []
    start = 1 if column is not None else 0
    selected = column if column is not None else first_non_empty_column(rows)
    for row in rows[start:]:
        if selected < len(row) and row[selected] is not None:
            value = str(row[selected]).strip()
            if value:
                values.append(value)
    return dedupe(values)


def preferred_column(header: list[object]) -> int | None:
    names = {"tag", "reference", "ref", "query", "value", "valeur"}
    for index, value in enumerate(header):
        if value is not None and str(value).strip().lower() in names:
            return index
    return None


def first_non_empty_column(rows: list[list[object]]) -> int:
    width = max((len(row) for row in rows), default=0)
    for column in range(width):
        if any(column < len(row) and row[column] not in (None, "") for row in rows):
            return column
    return 0


def dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        key = value.upper()
        if key not in seen:
            seen.add(key)
            output.append(value)
    return output


def run_batch_search(project_root: Path, references: list[str], mode: str = "hybrid", per_reference_limit: int = 25) -> list[BatchReferenceResult]:
    return [summarize_reference(reference, search(project_root, reference, mode, per_reference_limit)) for reference in references]


def summarize_reference(reference: str, results: list[SearchResult]) -> BatchReferenceResult:
    hits = [result for result in results if result.match_type != "not_found_in_indexed_text"]
    if not hits:
        return BatchReferenceResult(reference, "absent_indexed_text", 0, "", "", "", results[0].excerpt if results else "")
    documents = sorted({Path(result.document_path).name for result in hits if result.document_path})
    pages = sorted({str(result.page) for result in hits if result.page})
    best = hits[0]
    return BatchReferenceResult(
        reference=reference,
        status="found",
        hit_count=len(hits),
        documents=", ".join(documents),
        pages=", ".join(pages),
        best_match_type=best.match_type,
        best_excerpt=best.excerpt,
    )
